# 2022/03/22
# myself_webcam

# import threading
import openpyxl
import argparse
from sys import platform
import os
import sys
import cv2
import numpy as np

#-------------------------------------------------------------#
# Switch
SAVE_REC = True  # 是否將有姿態辨識過後的影片存檔在output_sample_videos
SAVE_EXCEL = False  # 是否儲存特徵點到output.xlsx
PREVIEW_INPUT_VIDEO_WITH_OPENPOSE_DETECT = False  # 是否預覽帶有姿態辨識過後的完整(無裁切)影片
#-------------------------------------------------------------#
# Input argument
signLanguageLabel = "snack"  # 鹹:salty 小吃:snack
dirPath = r'C:\Users\User\Desktop\snack'  # Input video的資料夾路徑
#-------------------------------------------------------------#

# my_answer = list()
my_answer = dict()
height = 0
width = 0
l_wrist = [200, 200, 200]
l_wrist = np.array(l_wrist)
r_wrist = [200, 200, 200]
r_wrist = np.array(r_wrist)
Recording = False

# catch keypoints:
my_keypoints_vectors = list()


# 列出指定路徑底下所有檔案(包含資料夾)
allFileList = os.listdir(dirPath)  # allFileList: 為所有input影片檔案名稱


def computeDistance(p1, p2):
    return ((p1[0]-p2[0])**2 + (p1[1]-p2[1])**2) ** 0.5


def write_xlsx(file_name, all_data):
    global signLanguageLabel
    # wksheet = wkbook["工作表1"]
    try:
        wkbook = openpyxl.load_workbook(file_name)
        # wkbook = openpyxl.Workbook(write_only=True)

    except Exception as e:
        # print(e)
        # 新增空白excel
        fn = file_name
        wkbook = openpyxl.Workbook()
        wkbook.save(fn)

    try:
        wksheet = wkbook["Sheet1"]
    except Exception as e:
        # print(e)
        # 新增空白sheet
        wksheet = wkbook.create_sheet("Sheet1", 0)
        wkbook.save(fn)

    wksheet.cell(row=1, column=1).value = "condition"
    begining = wksheet.max_row + 1
    i = 1

    wksheet.cell(row=begining, column=i).value = signLanguageLabel
    i += 1
    for data in all_data:
        try:
            wksheet.cell(row=begining, column=i).value = data[0]
            wksheet.cell(row=begining, column=i + 1).value = data[1]
            i = i + 2
        except Exception as e:
            print("資料格式錯誤")

    # wksheet.append(data.tolist())
    # for data in all_data:
    #     wksheet.append(data[0])
    #     wksheet.append(data[1])
    wkbook.save(file_name)

    # 關檔
    wb = openpyxl.Workbook(write_only=True)
    wb.close()


# def compute_vector(point1, point2):
#     return [point2[0] - point1[0], point2[1] - point1[1]]


def rescale_frame(frame, percent):
    wid = int(frame.shape[1] * percent / 100)
    hei = int(frame.shape[0] * percent / 100)
    dim = (wid, hei)
    return cv2.resize(frame, dim, interpolation=cv2.INTER_AREA)


def process_image(img, frame_num, opWrapper):
    global width
    global height
    global l_wrist, r_wrist

    # zoom_out = 25
    # l1, l2, l3, l4 = (width/2)-175, (height/2)-80, 350, 350
    # r1, r2, r3, r4 = (width/2)-175, (height/2)-80, 350, 350

    zoom_out = 100
    img = rescale_frame(img, percent=zoom_out)

    width_HandRectangle = 250*(zoom_out/100)

    l1, l2, l3, l4 = l_wrist[0]-width_HandRectangle/2, l_wrist[1] - \
        width_HandRectangle/2, width_HandRectangle, width_HandRectangle
    r1, r2, r3, r4 = r_wrist[0]-width_HandRectangle/2, r_wrist[1] - \
        width_HandRectangle/2, width_HandRectangle, width_HandRectangle

    handRectangles = [
        [
            op.Rectangle(l1, l2, l3, l4),
            op.Rectangle(r1, r2, r3, r4),
            # op.Rectangle(x, y, dx, dy),
            # op.Rectangle(88.984360, 268.866547,
            #              117.818230, 117.818230),
        ]
    ]

    # Create new datum
    datum = op.Datum()
    datum.cvInputData = img
    datum.handRectangles = handRectangles
    # datum.faceRectangles = faceRectangles

    # Process and display image
    opWrapper.emplaceAndPop(op.VectorDatum([datum]))

    # print(type(datum.cvOutputData))
    # img = np.zeros((512, 512, 3), np.uint8)
    # print(type(img))

    try:
        global Recording
        catch_poseKeypoints = datum.poseKeypoints.squeeze()
        catch_lefthandKeypoints = datum.handKeypoints[0].squeeze()
        catch_righthandKeypoints = datum.handKeypoints[1].squeeze()

        if(catch_poseKeypoints.shape != (25, 3)):
            return
        # print(catch_poseKeypoints.shape)
        # if (int(catch_poseKeypoints[7][1]-catch_poseKeypoints[1][1]) > 0 and int(catch_poseKeypoints[4][1]-catch_poseKeypoints[1][1]) > 0) or (catch_poseKeypoints[7][1] == 0 and catch_poseKeypoints[4][1] == 0):
        #     Recording = False
        # else:
        #     Recording = True
        try:
            rec_hori = computeDistance(
                catch_poseKeypoints[16], catch_poseKeypoints[17])
        except Exception as e:
            print(e)
            print(catch_poseKeypoints.__len__())
        # if (catch_poseKeypoints[1][1]+rec_hori) >= height:
        #     outOfRange = True
        # else:
        #     outOfRange = False
        if catch_poseKeypoints[16][1] == 0 or catch_poseKeypoints[17][1] == 0:
            tooClose = True
        else:
            tooClose = False

        if ((int(catch_poseKeypoints[7][1]-catch_poseKeypoints[1][1]) < int(0+rec_hori)) and
                catch_poseKeypoints[7][1] != 0) or ((int(catch_poseKeypoints[4][1]-catch_poseKeypoints[1][1]) < int(0+rec_hori)) and catch_poseKeypoints[4][1] != 0) and (not tooClose):
            Recording = True
        else:
            Recording = False

        # vector:
        #! 0->1 (pose)
        #! 1->2 2->3 3->4 (pose)
        # hand:
        # ? 0->1 1->2 2->3 3->4 | 0->5 5->6 6->7 7->8 | 0->9 9->10 10->11 11->12 | 0->13 13->14 14->15 15->16 |  0->17 17->18 18->19 19->20
        #! 1->5 5->6 6->7 (pose)
        # hand:
        # ? 0->1 1->2 2->3 3->4 | 0->5 5->6 6->7 7->8 | 0->9 9->10 10->11 11->12 | 0->13 13->14 14->15 15->16 |  0->17 17->18 18->19 19->20

        # * hand_sequence = [(0, 1), (1, 2), (2, 3), (3, 4),
        # *                  (0, 5), (5, 6), (6, 7), (7, 8),
        # *                  (0, 9), (9, 10), (10, 11), (11, 12),
        # *                  (0, 13), (13, 14), (14, 15), (15, 16),
        # *                 (0, 17), (17, 18), (18, 19), (19, 20)]
        # * pose_sequence = [(0, 1),
        # *                 (1, 2), (2, 3), (3, 4),
        # *                 (1, 5), (5, 6), (6, 7)]

        # my_keypoints_vector = list()

        # * for p1, p2 in pose_sequence:
        # *     my_keypoints_vectors.append(compute_vector(
        # *         catch_poseKeypoints[p1], catch_poseKeypoints[p2]))

        # * for p1, p2 in hand_sequence:
        # *     my_keypoints_vectors.append(compute_vector(
        # *         catch_lefthandKeypoints[p1], catch_lefthandKeypoints[p2]))
        # *     my_keypoints_vectors.append(compute_vector(
        # *         catch_righthandKeypoints[p1], catch_righthandKeypoints[p2]))
        # * my_keypoints_vectors.append(my_keypoints_vector)
        if(Recording == True):
            for point in catch_poseKeypoints:
                my_keypoints_vectors.append(point)
            for point in catch_lefthandKeypoints:
                my_keypoints_vectors.append(point)
            for point in catch_righthandKeypoints:
                my_keypoints_vectors.append(point)

        # 為了動態更新手部辨識範圍
        l_wrist = catch_poseKeypoints[4]
        r_wrist = catch_poseKeypoints[7]
        # print(l_wrist)
        # print(r_wrist)
    except Exception as e:
        # print(e)
        print("123456---")

    img = datum.cvOutputData
    cv2.rectangle(img, (int(l1), int(l2)),
                  (int(l1+l3), int(l2+l3)), (0, 255, 0), 2)
    cv2.rectangle(img, (int(r1), int(r2)),
                  (int(r1+r3), int(r2+r3)), (255, 0, 0), 2)
    # cv2.rectangle(img, (int(f1), int(f2)),
    #               (int(f1+f3), int(f2+f3)), (0, 255, 255), 2)
    # if (not outOfRange) or (not tooClose):
    if (not tooClose):
        cv2.line(img, (0, int(catch_poseKeypoints[1][1]+rec_hori)),
                 (int(width), int(catch_poseKeypoints[1][1]+rec_hori)), (255, 0, 0), 2)
    if(Recording == True):
        my_answer[frame_num] = img
        cv2.putText(img, "REC", (10, 40), cv2.FONT_HERSHEY_SIMPLEX,
                    1, (0, 0, 255), 3, cv2.LINE_AA)
    if PREVIEW_INPUT_VIDEO_WITH_OPENPOSE_DETECT:
        cv2.imshow("preview", img)


def work():
    global width
    global height
    global my_answer
    global my_keypoints_vectors

    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--video", default="C:\\Users\\User\\Desktop\\openpose1\\examples\\media\\004.mp4")

    args = parser.parse_known_args()

    params = dict()
    params["model_folder"] = "../../../models/"
    params["net_resolution"] = "160x96"
    params["hand_net_resolution"] = "352x352"
    # params["face_net_resolution"] = "256x256"
    params["hand"] = True
    params["hand_detector"] = 2
    # params["face"] = True
    # params["face_detector"] = 1
    params["body"] = 1

    for i in range(0, len(args[1])):
        curr_item = args[1][i]
        if i != len(args[1])-1:
            next_item = args[1][i+1]
        else:
            next_item = "1"
        if "--" in curr_item and "--" in next_item:
            key = curr_item.replace('-', '')
            if key not in params:
                params[key] = "1"
        elif "--" in curr_item and "--" not in next_item:
            key = curr_item.replace('-', '')
            if key not in params:
                params[key] = next_item

    # Starting Openpose
    opWrapper = op.WrapperPython()
    opWrapper.configure(params)
    opWrapper.start()

    # Process video
    # 新增放output影片的資料夾
    if SAVE_REC:
        try:
            os.mkdir("output_sample_videos")
        except:
            print("已經有output_sample_videos資料夾了")

    # cap = cv2.VideoCapture(args[0].video)
    file_counter = 1
    for my_file in allFileList:
        print(f"video: {file_counter} / {len(allFileList)}")
        file_counter += 1
        my_answer = dict()
        my_keypoints_vectors = list()
        cap = cv2.VideoCapture(dirPath + '\\' + my_file)

        frame_num = -1
        width = cap.get(cv2.CAP_PROP_FRAME_WIDTH)
        height = cap.get(cv2.CAP_PROP_FRAME_HEIGHT)

        if SAVE_REC:
            writer = cv2.VideoWriter(
                f'.\output_sample_videos\output_{my_file[:-4]}.avi', cv2.VideoWriter_fourcc(*'XVID'), 10.0, (int(width), int(height)))

        input_video_frame_num = 1
        input_video_total_frame_num = cap.get(cv2.CAP_PROP_FRAME_COUNT)
        while True:
            if (input_video_frame_num > input_video_total_frame_num):
                break
            try:
                frame_num += 1
                ret, imageToProcess = cap.read()

                cap.set(cv2.CAP_PROP_POS_FRAMES, input_video_frame_num)
                process_image(imageToProcess, frame_num, opWrapper)

                # threads.append(threading.Thread(target=process_image,
                #    args=(imageToProcess, frame_num, opWrapper)))
                # threads[frame_num].start()
                if cv2.waitKey(1) & 0xFF == ord('q'):
                    break
                # cv2.waitKey(50)
                input_video_frame_num += 10
            except Exception as e:
                print(e)
                break

        # 釋放攝影機
        cap.release()

        # 關閉所有 OpenCV 視窗
        cv2.destroyAllWindows()
        # for i in range(len(threads)):
        # threads[i].start()
        # for i in range(len(threads)):
        # threads[i].join()

        if SAVE_REC:
            for key in my_answer:
                if key == 0:
                    # 動態調整手部偵測範圍的第一張不算
                    continue
                writer.write(my_answer[key])
                # cv2.imshow("OpenPose 1.7.0 - Tutorial Python API", my_answer[key])
                # time.sleep(1)
                # cv2.waitKey(100)
        if SAVE_EXCEL:
            write_xlsx("output.xlsx", my_keypoints_vectors)
        if SAVE_REC:
            writer.release()


# <Main>
try:
    # Import Openpose (Windows/Ubuntu/OSX)
    dir_path = os.path.dirname(os.path.realpath(__file__))
    try:
        # Windows Import
        if platform == "win32":
            sys.path.append(dir_path + '/../../python/openpose/Release')
            os.environ['PATH'] = os.environ['PATH'] + ';' + dir_path + \
                '/../../x64/Release;' + dir_path + '/../../bin;'
            import pyopenpose as op
        else:
            sys.path.append('../../python')
            from openpose import pyopenpose as op
    except ImportError as e:
        print('Error: OpenPose library could not be found. Did you enable `BUILD_PYTHON` in CMake and have this Python script in the right folder?')
        raise e

    work()
except Exception as e:
    print(e)
    sys.exit(-1)
# </Main>
