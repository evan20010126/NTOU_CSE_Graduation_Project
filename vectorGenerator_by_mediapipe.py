import cv2
import mediapipe as mp
import os
import openpyxl
import numpy as np
from sklearn.feature_extraction import img_to_graph

#-------------------------------------------------------------#
# Switch
SAVE_REC = False  # 是否將有姿態辨識過後的影片存檔在output_sample_videos
SAVE_EXCEL = True  # 是否儲存特徵點到output.xlsx
PREVIEW_INPUT_VIDEO_WITH_OPENPOSE_DETECT = True  # 是否預覽帶有姿態辨識過後的完整(無裁切)影片
#-------------------------------------------------------------#
# Input argument
signLanguageLabel = "salty"  # 鹹:salty 小吃:snack
# Input video的資料夾路徑
dirPath = r'..\media\salty'
#-------------------------------------------------------------#

mp_drawing = mp.solutions.drawing_utils
mp_drawing_styles = mp.solutions.drawing_styles
mp_hands = mp.solutions.hands
mp_pose = mp.solutions.pose
all_keypoints = None
previous_hand = ""

# 將特徵點存入excel


def write_xlsx(file_name, all_data):
    print("Writing excel...")
    global signLanguageLabel
    # wksheet = wkbook["工作表1"]
    try:
        wkbook = openpyxl.load_workbook(file_name)
        # wkbook = openpyxl.Workbook(write_only=True)

    except Exception as e:
        # print(e)
        # 新增空白excel
        fn = file_name
        wkbook = openpyxl.Workbook()
        wkbook.save(fn)

    try:
        wksheet = wkbook["Sheet1"]
    except Exception as e:
        # print(e)
        # 新增空白sheet
        wksheet = wkbook.create_sheet("Sheet1", 0)
        wkbook.save(fn)

    wksheet.cell(row=1, column=1).value = "condition"
    begining = wksheet.max_row + 1
    i = 1

    wksheet.cell(row=begining, column=i).value = signLanguageLabel
    i += 1
    for data in all_data:
        try:
            wksheet.cell(row=begining, column=i).value = data
            i = i + 1
        except Exception as e:
            print("資料格式錯誤")
    # wksheet.append(data.tolist())
    # for data in all_data:
    #     wksheet.append(data[0])
    #     wksheet.append(data[1])
    wkbook.save(file_name)

    # 關檔
    wb = openpyxl.Workbook(write_only=True)
    wb.close()
    print("\Finish writing excel/")

# get landmarks


# def get_coordinates(Landmark, mode):
#     """
#     Get bounding box coordinates for a hand landmark.
#     Args:
#         handLadmark: A HandLandmark object.
#         image_shape: A tuple of the form (height, width).
#     Returns:
#         A tuple of the form (xmin, ymin, xmax, ymax).
#     """
#  # store all x and y points in list
#     if mode == 0:
#         for i in range(21):
#             # multiply x by image width (原本有*image_shape[1])
#             all_keypoints.append(Landmark.landmark[i].x)
#             # multiply y by image height (原本有*image_shape[0])
#             all_keypoints.append(Landmark.landmark[i].y)
#     if mode == 1:
#         for i in range(21):
#             # multiply x by image width (原本有*image_shape[1])
#             all_keypoints.append(Landmark.landmark[i].x)
#             # multiply y by image height (原本有*image_shape[0])
#             all_keypoints.append(Landmark.landmark[i].y)
#     return
first = True
miss_point = False


def computeDistance(p1, p2):
    return ((p1[0]-p2[0])**2 + (p1[1]-p2[1])**2) ** 0.5


def get_label_and_points(index, hand, results, hand_num):
    global previous_hand, first
    global all_keypoints
    global frame_keypoints_hands
    global image
    global catch_error
    # result.multi_handedness放此手的label跟score
    for idx, classification in enumerate(results.multi_handedness):
        # print(idx)
        # # print(classification.classification)
        # print(f"index: {index}")
        # if classification.classification[0].index == index:
        if idx == index:    #

            # Process results
            label = classification.classification[0].label
            score = classification.classification[0].score
            # text = '{} {}'.format(label, round(score, 2))

            # Extract Coordinates
            # coords = tuple(np.multiply(
            #     np.array((hand.landmark[mp_hands.HandLandmark.WRIST].x,
            #              hand.landmark[mp_hands.HandLandmark.WRIST].y)),
            #     [640, 480]).astype(int))
            # print("label: ", label)

            if(hand_num == 2):
                for i in range(21):
                    temp_xy = np.array(list())
                    temp_xy = np.append(
                        temp_xy, hand.landmark[i].x * image.shape[1])
                    temp_xy = np.append(
                        temp_xy, hand.landmark[i].y * image.shape[0])
                    frame_keypoints_hands = np.append(
                        frame_keypoints_hands, temp_xy)
            elif(hand_num == 1):
                if label == "Left":
                    for i in range(21):
                        temp_xy = np.array(list())
                        temp_xy = np.append(
                            temp_xy, hand.landmark[i].x * image.shape[1])
                        temp_xy = np.append(
                            temp_xy, hand.landmark[i].y * image.shape[0])
                        frame_keypoints_hands = np.append(
                            frame_keypoints_hands, temp_xy)
                    for i in range(21):
                        temp_xy = np.array([0, 0])
                        frame_keypoints_hands = np.append(
                            frame_keypoints_hands, temp_xy)
                elif label == "Right":
                    for i in range(21):
                        temp_xy = np.array([0, 0])
                        frame_keypoints_hands = np.append(
                            frame_keypoints_hands, temp_xy)
                    for i in range(21):
                        temp_xy = np.array(list())
                        temp_xy = np.append(
                            temp_xy, hand.landmark[i].x * image.shape[1])
                        temp_xy = np.append(
                            temp_xy, hand.landmark[i].y * image.shape[0])
                        frame_keypoints_hands = np.append(
                            frame_keypoints_hands, temp_xy)
            else:  # 有第三隻手
                catch_error = True
        # print("hand.landmark.size: ", len(hand.landmark)) # = 21
    # print(f"handnum{hand_num}")
    # print(frame_keypoints_hands)
    return


# <Main>
# 列出指定路徑底下所有檔案(包含資料夾)
allFileList = os.listdir(dirPath)  # allFileList: 為所有input影片檔案名稱

file_counter = 1
break_processing = False
frame_keypoints_pose = None
frame_keypoints_hands = None
image = None

for my_file in allFileList:
    if break_processing:
        break
    cap = cv2.VideoCapture(f"{dirPath}\\{my_file}")
    # cap = cv2.VideoCapture(0)
    print(f"video: {file_counter} / {len(allFileList)}")
    all_keypoints = list()
    file_counter += 1
    with mp_hands.Hands(
            model_complexity=0,
            min_detection_confidence=0.5,
            min_tracking_confidence=0.5) as hands, mp_pose.Pose(
            min_detection_confidence=0.5,
            min_tracking_confidence=0.5) as pose:
        while cap.isOpened():
            frame_keypoints_pose = np.array(list())
            frame_keypoints_hands = np.array(list())
            catch_error = False
            success, image = cap.read()
            if not success:
                print("Ignoring empty camera frame.")
                # If loading a video, use 'break' instead of 'continue'.
                break

            # To improve performance, optionally mark the image as not writeable to
            # pass by reference.
            image.flags.writeable = False
            image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
            results = hands.process(image)
            results_pose = pose.process(image)

            # Draw the hand annotations on the image.
            image.flags.writeable = True
            image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)

            mp_drawing.draw_landmarks(
                image,
                results_pose.pose_landmarks,
                mp_pose.POSE_CONNECTIONS,
                landmark_drawing_spec=mp_drawing_styles.get_default_pose_landmarks_style())
            # print("pose:")
            if results_pose.pose_landmarks:  # 當有偵測到pose
                # frame_keypoints_pose
                for i in range(23):  # 上半身的點(0~22)
                    # print(f"pose{i}")
                    # print(results_pose.pose_landmarks.landmark[i])
                    if (results_pose.pose_landmarks.landmark[i].visibility >= 0.5):
                        # ! 先隨便設，信心度超過0.5才填
                        temp_xy = np.array(list())
                        temp_xy = np.append(
                            temp_xy, results_pose.pose_landmarks.landmark[i].x * image.shape[1])
                        temp_xy = np.append(
                            temp_xy, results_pose.pose_landmarks.landmark[i].y * image.shape[0])

                        frame_keypoints_pose = np.append(
                            frame_keypoints_pose, temp_xy)
                        # all_keypoints.append(
                        #     results_pose.pose_landmarks.landmark[i].x * image.shape[1])
                        # all_keypoints.append(
                        #     results_pose.pose_landmarks.landmark[i].y * image.shape[0])
                    else:
                        #! 信心度太低就填0
                        temp_xy = np.array([0, 0])
                        frame_keypoints_pose = np.append(
                            frame_keypoints_pose, temp_xy)
                        # all_keypoints.append(0)
                        # all_keypoints.append(0)
                # print(results_pose.pose_landmarks.landmark[0].x)
            else:
                # for i in range(46):  # (23*2) = 46
                #     all_keypoints.append(0)
                continue

            if results.multi_hand_landmarks:
                # len(results.multi_hand_landmarks)代表有抓到幾隻手
                # <--畫手-->
                for num, hand_landmarks in enumerate(results.multi_hand_landmarks):
                    mp_drawing.draw_landmarks(
                        image,
                        hand_landmarks,
                        mp_hands.HAND_CONNECTIONS,
                        mp_drawing_styles.get_default_hand_landmarks_style(),
                        mp_drawing_styles.get_default_hand_connections_style())
                    get_label_and_points(num, hand_landmarks, results, len(
                        results.multi_hand_landmarks))
                    # print(all_keypoints)
                # <--End-->
            else:  # 找不到任何手
                catch_error = True

            frame_keypoints_pose = frame_keypoints_pose.reshape(-1, 2)
            frame_keypoints_hands = frame_keypoints_hands.reshape(-1, 2)
            Recording = False
            face_width = computeDistance(
                frame_keypoints_pose[7], frame_keypoints_pose[8])

            normalize_distance = computeDistance(
                frame_keypoints_pose[11], frame_keypoints_pose[12])
            normalize_original_point = (
                frame_keypoints_pose[12] + frame_keypoints_pose[11]) / 2

            # <Recording or not recording>
            if (frame_keypoints_pose[16][1] < normalize_original_point[1]+face_width and frame_keypoints_pose[16][1] != 0) or (frame_keypoints_pose[15][1] < normalize_original_point[1]+face_width and frame_keypoints_pose[15][1] != 0):
                Recording = True

            print(f"Recording: {Recording}")
            print(f"catch_error: {catch_error}")

            if Recording and not catch_error:
                # <Normalize>
                print("enter")
                for i in range(23):
                    all_keypoints.append(
                        (frame_keypoints_pose[i][0] - normalize_original_point[0])/normalize_distance)
                    all_keypoints.append(
                        (frame_keypoints_pose[i][1] - normalize_original_point[1])/normalize_distance)
                for i in range(42):
                    all_keypoints.append(
                        (frame_keypoints_hands[i][0] - normalize_original_point[0])/normalize_distance)
                    all_keypoints.append(
                        (frame_keypoints_hands[i][1] - normalize_original_point[1])/normalize_distance)
                cv2.putText(image, "REC", (10, 40), cv2.FONT_HERSHEY_SIMPLEX,
                            1, (0, 0, 255), 3, cv2.LINE_AA)

            # print("hands:")
            # print(results.multi_handedness)
            # if (results.multi_handedness != None):
            #     break_processing = True
            #     break

            # Flip the image horizoㄜntally for a selfie-view display.
            # cv2.imshow('MediaPipe Hands', cv2.flip(image, 1))
            cv2.imshow('MediaPipe Hands', image)

            if cv2.waitKey(5) & 0xFF == 27:
                break_processing = True
                break
    cap.release()
    if SAVE_EXCEL:
        write_xlsx("output.xlsx", all_keypoints)

print(all_keypoints)

# f = open("te", mode="w")
# f.write(all_keypoints.__str__())
# f.close()
