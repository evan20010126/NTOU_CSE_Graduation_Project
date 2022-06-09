from asyncio.windows_events import NULL
import openpyxl
from openpyxl import Workbook

file_name = "Summary_modify_edmund.xlsx"
wkbook = openpyxl.load_workbook(file_name)
wksheet = wkbook["Sheet1"]

# print(wksheet.max_row)
for i in range(2, wksheet.max_row+1):
    for j in range(2, wksheet.max_column+1):
        if wksheet.cell(row=i, column=j).value == None:
            wksheet.cell(row=i, column=j).value = 0
            print(i, " ", j)

wkbook.save("Summary_stuff_zero_4st.xlsx")
