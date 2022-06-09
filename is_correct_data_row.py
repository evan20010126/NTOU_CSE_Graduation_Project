from asyncio.windows_events import NULL
import openpyxl
from openpyxl import Workbook

file_name = "Summary.xlsx"
wkbook = openpyxl.load_workbook(file_name)
wksheet = wkbook["Sheet1"]

# print(wksheet.max_row)
row_max = 0
for i in range(2, wksheet.max_row+1):
    print("目前:" + str(i) + "行")
    for j in range(2, wksheet.max_column+1):
        if wksheet.cell(row=i, column=j).value == None:
            row_max = j - 1
            if (row_max - 1) % 134 != 0:  # 扣1是因為把class那欄拿掉
                print("row:" + str(i) + "，沒有整除，" +
                      "element數量:" + str(row_max-1))
            break
        if (j == wksheet.max_column):
            row_max = j
            if (row_max - 1) % 134 != 0:  # 扣1是因為把class那欄拿掉
                print("row:" + str(i) + "，沒有整除，" +
                      "element數量:" + str(row_max-1))
            break
