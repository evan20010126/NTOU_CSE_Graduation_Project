# 2022/03/22
# myself_webcam

# import threading
from nbformat import write
import openpyxl
import argparse
from sys import platform
import os
import sys
from turtle import width
import cv2
import numpy as np

from cv2 import CAP_PROP_FRAME_WIDTH
from cv2 import CAP_PROP_FRAME_HEIGHT

# my_answer = list()
my_answer = dict()
height = 0
width = 0
l_wrist = [200, 200, 200]
l_wrist = np.array(l_wrist)
r_wrist = [200, 200, 200]
r_wrist = np.array(r_wrist)

# catch keypoints:
my_keypoints_vectors = list()


def write_xlsx(file_name, all_data):
    wkbook = openpyxl.load_workbook(file_name)
    wksheet = wkbook["Sheet1"]
    wksheet.cell(row=1, column=1).value = "title"
    begining = wksheet.max_row + 1
    i = 1
    for data in all_data:
        wksheet.cell(row=begining, column=i).value = data[0]
        wksheet.cell(row=begining, column=i+1).value = data[1]
        i = i + 2

    # wksheet.append(data.tolist())
    # for data in all_data:
    #     wksheet.append(data[0])
    #     wksheet.append(data[1])
    wkbook.save(file_name)


def compute_vector(point1, point2):
    return [point2[0] - point1[0], point2[1] - point1[1]]


def process_image(img, frame_num, opWrapper):
    global width
    global height
    global l_wrist, r_wrist

    # zoom_out = 25
    # l1, l2, l3, l4 = (width/2)-175, (height/2)-80, 350, 350
    # r1, r2, r3, r4 = (width/2)-175, (height/2)-80, 350, 350

    l1, l2, l3, l4 = l_wrist[0]-150, l_wrist[1]-150, 300, 300
    r1, r2, r3, r4 = r_wrist[0]-150, r_wrist[1]-150, 300, 300

    handRectangles = [
        [
            op.Rectangle(l1, l2, l3, l4),
            op.Rectangle(r1, r2, r3, r4),
            # op.Rectangle(x, y, dx, dy),
            # op.Rectangle(88.984360, 268.866547,
            #              117.818230, 117.818230),
        ]
    ]

    # Create new datum
    datum = op.Datum()
    datum.cvInputData = img
    datum.handRectangles = handRectangles
    # datum.faceRectangles = faceRectangles

    # Process and display image
    opWrapper.emplaceAndPop(op.VectorDatum([datum]))

    # print(type(datum.cvOutputData))
    # img = np.zeros((512, 512, 3), np.uint8)
    # print(type(img))

    try:
        catch_poseKeypoints = datum.poseKeypoints.squeeze()
        catch_lefthandKeypoints = datum.handKeypoints[0].squeeze()
        catch_righthandKeypoints = datum.handKeypoints[1].squeeze()

        # vector:
        #! 0->1 (pose)
        #! 1->2 2->3 3->4 (pose)
        # hand:
        # ? 0->1 1->2 2->3 3->4 | 0->5 5->6 6->7 7->8 | 0->9 9->10 10->11 11->12 | 0->13 13->14 14->15 15->16 |  0->17 17->18 18->19 19->20
        #! 1->5 5->6 6->7 (pose)
        # hand:
        # ? 0->1 1->2 2->3 3->4 | 0->5 5->6 6->7 7->8 | 0->9 9->10 10->11 11->12 | 0->13 13->14 14->15 15->16 |  0->17 17->18 18->19 19->20

        # * hand_sequence = [(0, 1), (1, 2), (2, 3), (3, 4),
        # *                  (0, 5), (5, 6), (6, 7), (7, 8),
        # *                  (0, 9), (9, 10), (10, 11), (11, 12),
        # *                  (0, 13), (13, 14), (14, 15), (15, 16),
        # *                 (0, 17), (17, 18), (18, 19), (19, 20)]
        # * pose_sequence = [(0, 1),
        # *                 (1, 2), (2, 3), (3, 4),
        # *                 (1, 5), (5, 6), (6, 7)]

        # my_keypoints_vector = list()

        # * for p1, p2 in pose_sequence:
        # *     my_keypoints_vectors.append(compute_vector(
        # *         catch_poseKeypoints[p1], catch_poseKeypoints[p2]))

        # * for p1, p2 in hand_sequence:
        # *     my_keypoints_vectors.append(compute_vector(
        # *         catch_lefthandKeypoints[p1], catch_lefthandKeypoints[p2]))
        # *     my_keypoints_vectors.append(compute_vector(
        # *         catch_righthandKeypoints[p1], catch_righthandKeypoints[p2]))
        # * my_keypoints_vectors.append(my_keypoints_vector)
        for point in catch_poseKeypoints:
            my_keypoints_vectors.append(point)
        for point in catch_poseKeypoints:
            my_keypoints_vectors.append(point)
        for point in catch_poseKeypoints:
            my_keypoints_vectors.append(point)

        # 為了動態更新手部辨識範圍
        l_wrist = catch_poseKeypoints[4]
        r_wrist = catch_poseKeypoints[7]
        # print(l_wrist)
        # print(r_wrist)
    except Exception as e:
        # print(e)
        print("123456---")

    img = datum.cvOutputData
    cv2.rectangle(img, (int(l1), int(l2)),
                  (int(l1+l3), int(l2+l3)), (0, 255, 0), 2)
    cv2.rectangle(img, (int(r1), int(r2)),
                  (int(r1+r3), int(r2+r3)), (255, 0, 0), 2)
    # cv2.rectangle(img, (int(f1), int(f2)),
    #               (int(f1+f3), int(f2+f3)), (0, 255, 255), 2)
    my_answer[frame_num] = img
    cv2.imshow("preview", img)


def work():
    global width
    global height

    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--video", default="C:\\Users\\User\\Desktop\\openpose1\\examples\\media\\004.mp4")

    args = parser.parse_known_args()

    params = dict()
    params["model_folder"] = "../../../models/"
    params["net_resolution"] = "160x96"
    params["hand_net_resolution"] = "352x352"
    # params["face_net_resolution"] = "256x256"
    params["hand"] = True
    params["hand_detector"] = 2
    # params["face"] = True
    # params["face_detector"] = 1
    params["body"] = 1

    for i in range(0, len(args[1])):
        curr_item = args[1][i]
        if i != len(args[1])-1:
            next_item = args[1][i+1]
        else:
            next_item = "1"
        if "--" in curr_item and "--" in next_item:
            key = curr_item.replace('-', '')
            if key not in params:
                params[key] = "1"
        elif "--" in curr_item and "--" not in next_item:
            key = curr_item.replace('-', '')
            if key not in params:
                params[key] = next_item

    # Starting Openpose
    opWrapper = op.WrapperPython()
    opWrapper.configure(params)
    opWrapper.start()

    # Process video
    # cap = cv2.VideoCapture(args[0].video)
    cap = cv2.VideoCapture(0)

    frame_num = -1
    threads = []
    width = cap.get(CAP_PROP_FRAME_WIDTH)
    height = cap.get(CAP_PROP_FRAME_HEIGHT)
    writer = cv2.VideoWriter(
        '.\samplevideo.avi', cv2.VideoWriter_fourcc(*'XVID'), 10.0, (int(width), int(height)))

    while True:
        try:
            frame_num += 1
            ret, imageToProcess = cap.read()

            process_image(imageToProcess, frame_num, opWrapper)

            # threads.append(threading.Thread(target=process_image,
            #    args=(imageToProcess, frame_num, opWrapper)))
            # threads[frame_num].start()
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
            # cv2.waitKey(50)
        except:
            break

    # 釋放攝影機
    cap.release()

    # 關閉所有 OpenCV 視窗
    cv2.destroyAllWindows()
    # for i in range(len(threads)):
    # threads[i].start()
    # for i in range(len(threads)):
    # threads[i].join()

    for key in my_answer:
        if key == 0:
            # 動態調整手部偵測範圍的第一張不算
            continue
        writer.write(my_answer[key])
        # cv2.imshow("OpenPose 1.7.0 - Tutorial Python API", my_answer[key])
        # time.sleep(1)
        # cv2.waitKey(100)

    write_xlsx("output.xlsx", my_keypoints_vectors)
    writer.release()


# <Main>
try:
    # Import Openpose (Windows/Ubuntu/OSX)
    dir_path = os.path.dirname(os.path.realpath(__file__))
    try:
        # Windows Import
        if platform == "win32":
            sys.path.append(dir_path + '/../../python/openpose/Release')
            os.environ['PATH'] = os.environ['PATH'] + ';' + dir_path + \
                '/../../x64/Release;' + dir_path + '/../../bin;'
            import pyopenpose as op
        else:
            sys.path.append('../../python')
            from openpose import pyopenpose as op
    except ImportError as e:
        print('Error: OpenPose library could not be found. Did you enable `BUILD_PYTHON` in CMake and have this Python script in the right folder?')
        raise e

    work()
except Exception as e:
    print(e)
    sys.exit(-1)
# </Main>
